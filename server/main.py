"""
Sopra Steria Lynskarp — Central server (event-based)

Booth:  /e/{slug}/booth/{n}
Admin:  /admin  (React app) + /admin/api/...
"""

from __future__ import annotations

import asyncio
import csv as csv_mod
import io
import os
import random
import secrets
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import qrcode
from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from PIL import Image, ImageDraw, ImageFont

import db as instant_db
from event_defaults import default_event_fields, default_flow, normalize_event, slugify

_ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "soprasteria")
_security = HTTPBasic()


def require_admin(credentials: HTTPBasicCredentials = Depends(_security)):
    ok = secrets.compare_digest(credentials.password.encode(), _ADMIN_PASSWORD.encode())
    if not ok:
        raise HTTPException(status_code=401, detail="Unauthorized",
                            headers={"WWW-Authenticate": "Basic"})


# In-memory stores — populated from InstantDB on startup
events: List[dict] = []
users: List[dict] = []
booths: List[dict] = []
presentations: List[dict] = []


def _generate_short_code(event_id: str = "") -> str:
    chars = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    pool = [u for u in users if not event_id or u.get("event_id") == event_id]
    existing = {u.get("short_code") for u in pool}
    for _ in range(100):
        code = "".join(random.choices(chars, k=5))
        if code not in existing:
            return code
    return "".join(random.choices(chars, k=7))


def _event_by_id(event_id: str) -> Optional[dict]:
    return next((e for e in events if e.get("id") == event_id), None)


def _event_by_slug(slug: str) -> Optional[dict]:
    return next((e for e in events if e.get("slug") == slug), None)


def _require_event_slug(slug: str) -> dict:
    ev = _event_by_slug(slug)
    if not ev:
        raise HTTPException(status_code=404, detail="event not found")
    return ev


def _users_for(event_id: str) -> List[dict]:
    return [u for u in users if u.get("event_id") == event_id]


def _booths_for(event_id: str) -> List[dict]:
    return [b for b in booths if b.get("event_id") == event_id]


def _default_event() -> Optional[dict]:
    return _event_by_slug("lynskarp") or (events[0] if events else None)


def _public_event(ev: dict) -> dict:
    return {
        "id": ev["id"],
        "name": ev.get("name", ""),
        "slug": ev.get("slug", ""),
        "starts_at": ev.get("starts_at", ""),
        "ends_at": ev.get("ends_at", ""),
        "interests": ev.get("interests", []),
        "lookup_mode": ev.get("lookup_mode", "name"),
        "allow_walkup_registration": bool(ev.get("allow_walkup_registration")),
        "privacy_title": ev.get("privacy_title", ""),
        "privacy_bullets": ev.get("privacy_bullets", []),
        "privacy_checkbox_label": ev.get("privacy_checkbox_label", ""),
        "flow": ev.get("flow") or default_flow(),
        "max_interests": int(ev.get("max_interests") or 3),
    }


def _admin_event(ev: dict) -> dict:
    eid = ev["id"]
    return {
        **_public_event(ev),
        "created_at": ev.get("created_at", 0),
        "attendee_count": len(_users_for(eid)),
        "booth_count": len(_booths_for(eid)),
    }


async def _ensure_short_codes():
    loop = asyncio.get_event_loop()
    count = 0
    for u in users:
        if not u.get("short_code"):
            u["short_code"] = _generate_short_code(u.get("event_id", ""))
            try:
                await loop.run_in_executor(
                    None, lambda uid=u["id"], sc=u["short_code"]: instant_db.update_user(uid, short_code=sc),
                )
            except Exception as e:
                print(f"[startup] Failed to persist short_code for {u['id']}: {e}")
            count += 1
    if count:
        print(f"[startup] Generated and persisted {count} short codes")


async def _migrate_to_events(loop):
    """Create default event and backfill event_id on legacy rows."""
    if not events:
        event_id = str(uuid.uuid4())
        fields = default_event_fields("Momentum Lynskarp", "lynskarp")
        ev = {"id": event_id, **fields, "created_at": int(time.time() * 1000)}
        ev = normalize_event(ev)
        events.append(ev)
        try:
            await loop.run_in_executor(None, instant_db.create_event, event_id, fields)
            print(f"[startup] Created default event Momentum Lynskarp ({event_id})")
        except Exception as e:
            print(f"[startup] Could not persist default event: {e}")
    else:
        for i, ev in enumerate(list(events)):
            events[i] = normalize_event(ev)

    default = _default_event()
    if not default:
        return
    eid = default["id"]

    for u in users:
        if not u.get("event_id"):
            u["event_id"] = eid
            try:
                await loop.run_in_executor(
                    None, lambda uid=u["id"]: instant_db.update_user(uid, event_id=eid),
                )
            except Exception as e:
                print(f"[startup] Failed to backfill user event_id: {e}")

    for b in booths:
        if not b.get("event_id"):
            b["event_id"] = eid
            try:
                await loop.run_in_executor(
                    None, lambda bid=b["id"]: instant_db.update_booth(bid, event_id=eid),
                )
            except Exception as e:
                print(f"[startup] Failed to backfill booth event_id: {e}")

    for p in presentations:
        if not p.get("event_id"):
            p["event_id"] = eid
            try:
                await loop.run_in_executor(
                    None, lambda pid=p["id"]: _patch_presentation_event(pid, eid),
                )
            except Exception as e:
                print(f"[startup] Failed to backfill presentation event_id: {e}")


def _patch_presentation_event(pres_id: str, event_id: str) -> None:
    import httpx
    payload = {"steps": [["update", "presentations", pres_id, {"event_id": event_id}]]}
    r = httpx.post(
        f"{instant_db._BASE}/admin/transact",
        json=payload,
        headers=instant_db._headers(),
        timeout=10,
    )
    r.raise_for_status()


@asynccontextmanager
async def lifespan(app: FastAPI):
    loop = asyncio.get_event_loop()
    try:
        evs = await loop.run_in_executor(None, instant_db.get_all_events)
        events.extend([normalize_event(e) for e in evs])
        print(f"[startup] Restored {len(evs)} events from InstantDB")
    except Exception as e:
        print(f"[startup] Could not restore events from InstantDB: {e}")

    try:
        us = await loop.run_in_executor(None, instant_db.get_all_users)
        users.extend(us)
        print(f"[startup] Restored {len(us)} users from InstantDB")
    except Exception as e:
        print(f"[startup] Could not restore users from InstantDB: {e}")

    try:
        bs = await loop.run_in_executor(None, instant_db.get_all_booths)
        booths.extend(bs)
        print(f"[startup] Restored {len(booths)} booths from InstantDB")
    except Exception as e:
        print(f"[startup] Could not restore booths from InstantDB: {e}")

    try:
        ps = await loop.run_in_executor(None, instant_db.get_all_presentations)
        presentations.extend(ps)
        print(f"[startup] Restored {len(ps)} presentations from InstantDB")
    except Exception as e:
        print(f"[startup] Could not restore presentations from InstantDB: {e}")

    await _migrate_to_events(loop)
    await _ensure_short_codes()

    # Ensure at least one booth on default event
    default = _default_event()
    if default and not _booths_for(default["id"]):
        booth_id = str(uuid.uuid4())
        booth = {"id": booth_id, "name": "Booth 1", "number": 1, "mode": "both", "event_id": default["id"]}
        booths.append(booth)
        try:
            await loop.run_in_executor(
                None, instant_db.create_booth, booth_id, "Booth 1", 1, "both", default["id"],
            )
            print("[startup] Created default booth 1 for lynskarp")
        except Exception as e:
            print(f"[startup] Could not create default booth: {e}")

    yield


app = FastAPI(lifespan=lifespan)


# --- Demo signup page (root) ---

@app.get("/", response_class=HTMLResponse)
def demo_signup_page():
    return (Path(__file__).parent / "static" / "demo-signup.html").read_text()


@app.get("/health")
def health():
    return {"ok": True, "users": len(users), "events": len(events)}


# --- Label generation (unchanged visually) ---

_QR_BASE_URL = os.getenv("QR_BASE_URL", "https://lynskarp.soprasteria.no")
_LABEL_W = round(103 * 300 / 25.4)
_LABEL_H = round(60 * 300 / 25.4)
_ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
_BUNDLED_FONT = str(_ASSETS_DIR / "DejaVuSans-Bold.ttf")


def _find_font(size):
    return ImageFont.truetype(_BUNDLED_FONT, size)


def _generate_label(user_name: str, interest: str, short_code: str) -> bytes:
    PAD = 14
    canvas = Image.new("RGB", (_LABEL_W, _LABEL_H), "white")
    draw = ImageDraw.Draw(canvas)

    NAME_PAD_TOP = 20
    name_font_size = 72
    name_font = _find_font(name_font_size)
    while name_font_size > 16:
        name_font = _find_font(name_font_size)
        bbox = name_font.getbbox(user_name)
        if (bbox[2] - bbox[0]) <= _LABEL_W - PAD * 2:
            break
        name_font_size -= 4
    name_bbox = name_font.getbbox(user_name)
    name_text_w = name_bbox[2] - name_bbox[0]
    name_text_h = name_bbox[3] - name_bbox[1]
    name_x = (_LABEL_W - name_text_w) // 2 - name_bbox[0]
    name_y = NAME_PAD_TOP - name_bbox[1]
    draw.text((name_x, name_y), user_name, fill="black", font=name_font)

    sep_y = NAME_PAD_TOP + name_text_h + 8
    draw.line([(PAD, sep_y), (_LABEL_W - PAD, sep_y)], fill="#cccccc", width=2)

    bottom_top = sep_y + 10
    bottom_h = _LABEL_H - bottom_top - PAD

    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=2)
    qr.add_data(short_code or "NOCODE")
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qr_size = bottom_h
    qr_img = qr_img.resize((qr_size, qr_size), Image.NEAREST)
    canvas.paste(qr_img, (PAD, bottom_top))

    text_left = PAD + qr_size + PAD * 2
    text_area_w = _LABEL_W - text_left - PAD * 2
    items = [s.strip() for s in (interest or "").split(",") if s.strip()]
    item_spacing = 35

    if items:
        interest_font_size = 60
        wrapped_lines = []
        while interest_font_size > 10:
            interest_font = _find_font(interest_font_size)
            line_h = interest_font.getbbox("Ag")[3] - interest_font.getbbox("Ag")[1]
            wrapped_lines = []
            words_fit = True
            for item in items:
                words = item.split()
                for word in words:
                    ww = interest_font.getbbox(word)[2] - interest_font.getbbox(word)[0]
                    if ww > text_area_w:
                        words_fit = False
                        break
                if not words_fit:
                    break
                lines = []
                current = words[0]
                for word in words[1:]:
                    test = current + " " + word
                    tw = interest_font.getbbox(test)[2] - interest_font.getbbox(test)[0]
                    if tw <= text_area_w:
                        current = test
                    else:
                        lines.append(current)
                        current = word
                lines.append(current)
                wrapped_lines.append(lines)
            if not words_fit:
                interest_font_size -= 4
                continue
            total_lines = sum(len(lines) for lines in wrapped_lines)
            total_h = total_lines * line_h + (len(items) - 1) * item_spacing
            if total_h <= bottom_h:
                break
            interest_font_size -= 4
        interest_font = _find_font(interest_font_size)
        line_h = interest_font.getbbox("Ag")[3] - interest_font.getbbox("Ag")[1]
        total_lines = sum(len(lines) for lines in wrapped_lines)
        total_h = total_lines * line_h + (len(items) - 1) * item_spacing
        y_cursor = bottom_top + (bottom_h - total_h) // 2
        for idx, lines in enumerate(wrapped_lines):
            for line in lines:
                bbox = interest_font.getbbox(line)
                draw.text((text_left, y_cursor - bbox[1]), line, fill="#444444", font=interest_font)
                y_cursor += line_h
            if idx < len(wrapped_lines) - 1:
                y_cursor += item_spacing

    buf = io.BytesIO()
    canvas.save(buf, format="PNG")
    return buf.getvalue()


def _generate_qr_png(data: str) -> bytes:
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=2)
    qr.add_data(data or "NOCODE")
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    buf = io.BytesIO()
    qr_img.save(buf, format="PNG")
    return buf.getvalue()


# ── Event-scoped booth / public API ───────────────────────────────────────────

@app.get("/e/{slug}/config")
def event_config(slug: str):
    ev = _require_event_slug(slug)
    return _public_event(ev)


@app.get("/e/{slug}/booth-config/{number}")
def event_booth_config(slug: str, number: int):
    ev = _require_event_slug(slug)
    booth = next((b for b in _booths_for(ev["id"]) if b.get("number") == number), None)
    return {"mode": booth.get("mode", "both") if booth else "both", "event": _public_event(ev)}


@app.get("/e/{slug}/users")
def event_search_users(slug: str, q: str = ""):
    ev = _require_event_slug(slug)
    q_lower = q.strip().lower()
    mode = ev.get("lookup_mode", "name")

    def _user_dict(u):
        return {
            "id": u["id"],
            "name": u.get("name", ""),
            "email": u.get("email", ""),
            "phone": u.get("phone", ""),
            "has_char": False,
        }

    pool = _users_for(ev["id"])
    if not q_lower:
        return [_user_dict(u) for u in pool]

    result = []
    for u in pool:
        name = u.get("name", "").lower()
        phone = (u.get("phone") or "").lower().replace(" ", "")
        qn = q_lower.replace(" ", "")
        ok = False
        if mode in ("name", "both") and q_lower in name:
            ok = True
        if mode in ("phone", "both") and qn and qn in phone:
            ok = True
        if ok:
            result.append(_user_dict(u))
    return result


@app.post("/e/{slug}/users")
async def event_create_walkup_user(slug: str, body: dict):
    """Walk-up registration when allow_walkup_registration is enabled."""
    ev = _require_event_slug(slug)
    if not ev.get("allow_walkup_registration"):
        raise HTTPException(status_code=403, detail="walk-up registration disabled")
    name = (body.get("name") or "").strip()
    email = (body.get("email") or "").strip()
    phone = (body.get("phone") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    eid = ev["id"]
    pool = _users_for(eid)
    if email and any(u.get("email", "").lower() == email.lower() for u in pool):
        raise HTTPException(status_code=409, detail="email already exists")
    if phone and any((u.get("phone") or "").replace(" ", "") == phone.replace(" ", "") for u in pool):
        raise HTTPException(status_code=409, detail="phone already exists")
    user_id = str(uuid.uuid4())
    user = {
        "id": user_id,
        "name": name,
        "email": email,
        "phone": phone,
        "event_id": eid,
        "created_at": int(time.time() * 1000),
        "short_code": _generate_short_code(eid),
    }
    users.append(user)
    try:
        await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: instant_db.create_user(
                user_id, name, email, user["short_code"], eid, phone,
            ),
        )
    except Exception as e:
        print(f"[walkup] create failed: {e}")
    return {"ok": True, "id": user_id, "name": name, "email": email, "phone": phone}


@app.get("/e/{slug}/users/{user_id}")
def event_get_user(slug: str, user_id: str):
    ev = _require_event_slug(slug)
    user = next((u for u in _users_for(ev["id"]) if u["id"] == user_id), None)
    if not user:
        raise HTTPException(status_code=404, detail="user not found")
    return {
        "id": user["id"],
        "name": user.get("name", ""),
        "interest": user.get("interest", ""),
        "short_code": user.get("short_code", ""),
        "wants_demo": bool(user.get("wants_demo")),
        "phone": user.get("phone", ""),
    }


@app.get("/e/{slug}/users/by-code/{code}")
def event_get_user_by_code(slug: str, code: str):
    ev = _require_event_slug(slug)
    code_upper = code.strip().upper()
    user = next((u for u in _users_for(ev["id"]) if u.get("short_code") == code_upper), None)
    if not user:
        raise HTTPException(status_code=404, detail="user not found")
    return {
        "id": user["id"],
        "name": user.get("name", ""),
        "interest": user.get("interest", ""),
        "wants_demo": bool(user.get("wants_demo")),
    }


@app.post("/e/{slug}/print-label")
async def event_print_label(
    slug: str,
    name: str = Form(""),
    interest: str = Form(""),
    user_id: str = Form(""),
):
    ev = _require_event_slug(slug)
    if not user_id:
        return {"ok": False, "error": "user_id required"}
    user = next((u for u in _users_for(ev["id"]) if u["id"] == user_id), None)
    if not user:
        return {"ok": False, "error": "user not found"}
    user["interest"] = interest
    user["label_printed"] = False
    short_code = user.get("short_code", "")
    try:
        await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: instant_db.update_user(
                user_id, interest=interest, label_printed=False, short_code=short_code,
            ),
        )
    except Exception as e:
        print(f"[print-label] DB write failed: {e}")
        return {"ok": False, "error": str(e)}
    print(f"[print-label] queued label for {name!r} event={slug} interest={interest!r}")
    return {"ok": True}


@app.post("/e/{slug}/demo-choice")
async def event_demo_choice(slug: str, body: dict):
    ev = _require_event_slug(slug)
    user_id = body.get("user_id")
    wants_demo = body.get("wants_demo", True)
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id required")
    user = next((u for u in _users_for(ev["id"]) if u["id"] == user_id), None)
    if not user:
        raise HTTPException(status_code=404, detail="user not found")
    user["wants_demo"] = bool(wants_demo)
    try:
        await asyncio.get_event_loop().run_in_executor(
            None, lambda: instant_db.update_user(user_id, wants_demo=bool(wants_demo)),
        )
    except Exception as e:
        print(f"[demo] DB write failed: {e}")
    return {"ok": True}


@app.get("/e/{slug}/label-preview/{user_id}")
def event_label_preview(slug: str, user_id: str, name: str = "", interest: str = ""):
    ev = _require_event_slug(slug)
    user = next((u for u in _users_for(ev["id"]) if u["id"] == user_id), None)
    if not user:
        raise HTTPException(status_code=404, detail="user not found")
    if not name:
        name = user.get("name", "")
        interest = interest or user.get("interest", "")
    png = _generate_label(name, interest, user.get("short_code", ""))
    return Response(content=png, media_type="image/png")


# ── Compatibility shims (legacy single-event routes) ──────────────────────────

def _legacy_slug() -> str:
    d = _default_event()
    return d["slug"] if d else "lynskarp"


@app.post("/print-label")
async def print_label_legacy(name: str = Form(""), interest: str = Form(""), user_id: str = Form("")):
    return await event_print_label(_legacy_slug(), name=name, interest=interest, user_id=user_id)


@app.get("/users")
def search_users_legacy(q: str = ""):
    return event_search_users(_legacy_slug(), q=q)


@app.get("/users/by-code/{code}")
def get_user_by_code_legacy(code: str):
    return event_get_user_by_code(_legacy_slug(), code)


@app.get("/users/{user_id}")
def get_user_legacy(user_id: str):
    return event_get_user(_legacy_slug(), user_id)


@app.post("/demo-choice")
async def demo_choice_legacy(body: dict):
    return await event_demo_choice(_legacy_slug(), body)


@app.get("/booth-config/{number}")
def booth_config_legacy(number: int):
    return event_booth_config(_legacy_slug(), number)


@app.get("/label-preview/{user_id}")
def label_preview_legacy(user_id: str, name: str = "", interest: str = ""):
    return event_label_preview(_legacy_slug(), user_id, name=name, interest=interest)


# ── Admin: Events ─────────────────────────────────────────────────────────────

@app.get("/admin/api/events")
def admin_list_events(_=Depends(require_admin)):
    return [_admin_event(e) for e in events]


@app.post("/admin/api/events")
async def admin_create_event(body: dict, _=Depends(require_admin)):
    name = (body.get("name") or "").strip() or "Nytt arrangement"
    slug = slugify(body.get("slug") or name)
    if any(e.get("slug") == slug for e in events):
        raise HTTPException(status_code=409, detail="slug already exists")
    fields = default_event_fields(name, slug)
    for key in (
        "starts_at", "ends_at", "interests", "lookup_mode", "allow_walkup_registration",
        "privacy_title", "privacy_bullets", "privacy_checkbox_label", "flow", "max_interests",
    ):
        if key in body and body[key] is not None:
            fields[key] = body[key]
    event_id = str(uuid.uuid4())
    ev = normalize_event({"id": event_id, **fields, "created_at": int(time.time() * 1000)})
    events.append(ev)
    try:
        await asyncio.get_event_loop().run_in_executor(
            None, instant_db.create_event, event_id, {k: ev[k] for k in fields},
        )
    except Exception as e:
        print(f"[admin] create event failed: {e}")

    # Default booth
    booth_id = str(uuid.uuid4())
    booth = {"id": booth_id, "name": "Booth 1", "number": 1, "mode": "both", "event_id": event_id}
    booths.append(booth)
    try:
        await asyncio.get_event_loop().run_in_executor(
            None, instant_db.create_booth, booth_id, "Booth 1", 1, "both", event_id,
        )
    except Exception as e:
        print(f"[admin] create default booth failed: {e}")

    return {"ok": True, **_admin_event(ev)}


@app.get("/admin/api/events/{event_id}")
def admin_get_event(event_id: str, _=Depends(require_admin)):
    ev = _event_by_id(event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="not found")
    return _admin_event(ev)


@app.patch("/admin/api/events/{event_id}")
async def admin_patch_event(event_id: str, body: dict, _=Depends(require_admin)):
    ev = _event_by_id(event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="not found")
    allowed = {
        "name", "slug", "starts_at", "ends_at", "interests", "lookup_mode",
        "allow_walkup_registration", "privacy_title", "privacy_bullets",
        "privacy_checkbox_label", "flow", "max_interests",
    }
    updates = {}
    for key in allowed:
        if key in body:
            updates[key] = body[key]
    if "slug" in updates:
        updates["slug"] = slugify(updates["slug"])
        if any(e.get("slug") == updates["slug"] and e["id"] != event_id for e in events):
            raise HTTPException(status_code=409, detail="slug already exists")
    if "lookup_mode" in updates and updates["lookup_mode"] not in ("name", "phone", "both"):
        raise HTTPException(status_code=400, detail="invalid lookup_mode")
    ev.update(updates)
    normalized = normalize_event(ev)
    idx = next(i for i, e in enumerate(events) if e["id"] == event_id)
    events[idx] = normalized
    try:
        await asyncio.get_event_loop().run_in_executor(
            None, lambda: instant_db.update_event(event_id, **updates),
        )
    except Exception as e:
        print(f"[admin] patch event failed: {e}")
    return {"ok": True, **_admin_event(normalized)}


@app.delete("/admin/api/events/{event_id}")
async def admin_delete_event(event_id: str, _=Depends(require_admin)):
    ev = _event_by_id(event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="not found")
    if len(events) <= 1:
        raise HTTPException(status_code=400, detail="cannot delete the last event")
    # Delete attendees and booths for event
    for u in list(_users_for(event_id)):
        users.remove(u)
        try:
            await asyncio.get_event_loop().run_in_executor(None, instant_db.delete_user, u["id"])
        except Exception as e:
            print(f"[admin] delete user failed: {e}")
    for b in list(_booths_for(event_id)):
        booths.remove(b)
        try:
            await asyncio.get_event_loop().run_in_executor(None, instant_db.delete_booth, b["id"])
        except Exception as e:
            print(f"[admin] delete booth failed: {e}")
    events.remove(ev)
    try:
        await asyncio.get_event_loop().run_in_executor(None, instant_db.delete_event, event_id)
    except Exception as e:
        print(f"[admin] delete event failed: {e}")
    return {"ok": True}


# ── Admin: Users (event-scoped) ───────────────────────────────────────────────

def _parse_spreadsheet(raw: bytes, filename: str) -> list[dict]:
    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            raise HTTPException(status_code=400, detail="Empty spreadsheet")
        headers = [str(h or "").strip().lower() for h in rows[0]]
        return [{headers[i]: (str(cell) if cell is not None else "") for i, cell in enumerate(row)}
                for row in rows[1:]]
    content = raw.decode("utf-8-sig")
    reader = csv_mod.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="Empty CSV or no header row")
    return [{f.strip().lower(): (row.get(f) or "").strip() for f in reader.fieldnames}
            for row in reader]


@app.get("/admin/api/events/{event_id}/users")
def admin_list_event_users(event_id: str, _=Depends(require_admin)):
    if not _event_by_id(event_id):
        raise HTTPException(status_code=404, detail="event not found")
    result = []
    for u in _users_for(event_id):
        result.append({
            "id": u["id"],
            "name": u.get("name", ""),
            "email": u.get("email", ""),
            "phone": u.get("phone", ""),
            "interest": u.get("interest", ""),
            "short_code": u.get("short_code", ""),
            "label_printed": u.get("label_printed"),
            "wants_demo": bool(u.get("wants_demo")),
            "created_at": u.get("created_at", 0),
            "event_id": event_id,
        })
    result.sort(key=lambda u: u.get("name", "").lower())
    return result


@app.post("/admin/api/events/{event_id}/users")
async def admin_add_event_user(event_id: str, body: dict, _=Depends(require_admin)):
    if not _event_by_id(event_id):
        raise HTTPException(status_code=404, detail="event not found")
    name = (body.get("name") or "").strip()
    email = (body.get("email") or "").strip()
    phone = (body.get("phone") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    pool = _users_for(event_id)
    if email and any(u.get("email", "").lower() == email.lower() for u in pool):
        raise HTTPException(status_code=409, detail="email already exists")
    user_id = str(uuid.uuid4())
    user = {
        "id": user_id, "name": name, "email": email, "phone": phone,
        "event_id": event_id, "created_at": int(time.time() * 1000),
        "short_code": _generate_short_code(event_id),
    }
    users.append(user)
    try:
        await asyncio.get_event_loop().run_in_executor(
            None, lambda: instant_db.create_user(user_id, name, email, user["short_code"], event_id, phone),
        )
    except Exception as e:
        print(f"[admin] create user failed: {e}")
    return {"ok": True, **user}


@app.post("/admin/api/events/{event_id}/import-users")
async def admin_import_event_users(event_id: str, file: UploadFile = File(...), _=Depends(require_admin)):
    if not _event_by_id(event_id):
        raise HTTPException(status_code=404, detail="event not found")
    raw = await file.read()
    rows = _parse_spreadsheet(raw, file.filename or "")
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found")
    col_keys = list(rows[0].keys())
    name_col = next((k for k in col_keys if "name" in k or k == "navn"), None)
    email_col = next((k for k in col_keys if "email" in k or "epost" in k), None)
    phone_col = next((k for k in col_keys if "phone" in k or "telefon" in k or "tlf" in k), None)
    if not name_col:
        raise HTTPException(status_code=400, detail=f"File must have Name column. Found: {col_keys}")
    pool = _users_for(event_id)
    existing_emails = {u.get("email", "").lower() for u in pool if u.get("email")}
    imported, skipped = 0, 0
    for row in rows:
        name = (row.get(name_col) or "").strip()
        email = (row.get(email_col) or "").strip() if email_col else ""
        phone = (row.get(phone_col) or "").strip() if phone_col else ""
        if not name:
            skipped += 1
            continue
        if email and email.lower() in existing_emails:
            skipped += 1
            continue
        user_id = str(uuid.uuid4())
        user = {
            "id": user_id, "name": name, "email": email, "phone": phone,
            "event_id": event_id, "created_at": int(time.time() * 1000),
            "short_code": _generate_short_code(event_id),
        }
        users.append(user)
        if email:
            existing_emails.add(email.lower())
        try:
            await asyncio.get_event_loop().run_in_executor(
                None,
                lambda uid=user_id, n=name, e=email, p=phone, sc=user["short_code"]:
                    instant_db.create_user(uid, n, e, sc, event_id, p),
            )
        except Exception as e:
            print(f"[import] DB write failed: {e}")
        imported += 1
    return {"ok": True, "imported": imported, "skipped": skipped}


@app.patch("/admin/api/events/{event_id}/users/{user_id}")
async def admin_patch_event_user(event_id: str, user_id: str, body: dict, _=Depends(require_admin)):
    user = next((u for u in _users_for(event_id) if u["id"] == user_id), None)
    if not user:
        raise HTTPException(status_code=404, detail="not found")
    db_updates = {}
    if body.get("clear_interest"):
        user.pop("interest", None)
        user.pop("label_printed", None)
        db_updates["interest"] = None
        db_updates["label_printed"] = None
    if "interest" in body and not body.get("clear_interest"):
        user["interest"] = body["interest"]
        db_updates["interest"] = body["interest"]
    if "name" in body:
        name = (body.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="name required")
        user["name"] = name
        db_updates["name"] = name
    if "email" in body:
        email = (body.get("email") or "").strip()
        user["email"] = email
        db_updates["email"] = email
    if "phone" in body:
        phone = (body.get("phone") or "").strip()
        user["phone"] = phone
        db_updates["phone"] = phone
    if body.get("reprint"):
        user["label_printed"] = False
        db_updates["label_printed"] = False
    if "wants_demo" in body:
        user["wants_demo"] = bool(body["wants_demo"])
        db_updates["wants_demo"] = bool(body["wants_demo"])
    if db_updates:
        try:
            await asyncio.get_event_loop().run_in_executor(
                None, lambda: instant_db.update_user(user_id, **db_updates),
            )
        except Exception as e:
            print(f"[admin] patch user failed: {e}")
    return {"ok": True}


@app.delete("/admin/api/events/{event_id}/users/{user_id}")
async def admin_delete_event_user(event_id: str, user_id: str, _=Depends(require_admin)):
    user = next((u for u in _users_for(event_id) if u["id"] == user_id), None)
    if not user:
        raise HTTPException(status_code=404, detail="not found")
    users.remove(user)
    try:
        await asyncio.get_event_loop().run_in_executor(None, instant_db.delete_user, user_id)
    except Exception as e:
        print(f"[admin] delete user failed: {e}")
    return {"ok": True}


@app.get("/admin/api/events/{event_id}/users/{user_id}/qr")
def admin_event_user_qr(event_id: str, user_id: str, _=Depends(require_admin)):
    user = next((u for u in _users_for(event_id) if u["id"] == user_id), None)
    if not user or not user.get("short_code"):
        raise HTTPException(status_code=404, detail="not found")
    return Response(content=_generate_qr_png(user["short_code"]), media_type="image/png")


@app.post("/admin/api/events/{event_id}/clear-registrations")
async def admin_clear_event_registrations(event_id: str, _=Depends(require_admin)):
    count = 0
    for u in _users_for(event_id):
        if u.get("interest") or u.get("wants_demo") or u.get("label_printed"):
            u.pop("interest", None)
            u.pop("label_printed", None)
            u["wants_demo"] = False
            try:
                await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda uid=u["id"]: instant_db.update_user(
                        uid, interest=None, label_printed=None, wants_demo=False,
                    ),
                )
            except Exception as e:
                print(f"[admin] clear failed: {e}")
            count += 1
    return {"ok": True, "cleared": count}


@app.post("/admin/api/events/{event_id}/delete-all-users")
async def admin_delete_all_event_users(event_id: str, _=Depends(require_admin)):
    pool = list(_users_for(event_id))
    for u in pool:
        users.remove(u)
        try:
            await asyncio.get_event_loop().run_in_executor(None, instant_db.delete_user, u["id"])
        except Exception as e:
            print(f"[admin] delete failed: {e}")
    return {"ok": True, "deleted": len(pool)}


@app.get("/admin/api/events/{event_id}/export-users")
def admin_export_event_users(event_id: str, _=Depends(require_admin)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    ev = _event_by_id(event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="not found")
    pool = _users_for(event_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Brukere"
    ws.append([f"Brukeroversikt — {ev.get('name', '')}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"{len(pool)} brukere totalt"])
    ws.append([])
    headers = ["Navn", "E-post", "Telefon", "Interesser", "Vil ha demo", "Etikett skrevet ut", "Kort-kode", "Registrert"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E0F0", end_color="E8E0F0", fill_type="solid")
    for u in sorted(pool, key=lambda x: x.get("name", "").lower()):
        created = u.get("created_at", 0)
        created_str = datetime.fromtimestamp(created / 1000).strftime("%Y-%m-%d %H:%M") if created else ""
        ws.append([
            u.get("name", ""), u.get("email", ""), u.get("phone", ""), u.get("interest", ""),
            "Ja" if u.get("wants_demo") else "Nei",
            "Ja" if u.get("label_printed") else "Nei",
            u.get("short_code", ""), created_str,
        ])
    for col, w in zip("ABCDEFGH", [25, 30, 16, 40, 14, 20, 12, 18]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=brukeroversikt.xlsx"},
    )


@app.get("/admin/api/events/{event_id}/export-interests")
def admin_export_event_interests(event_id: str, _=Depends(require_admin)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    ev = _event_by_id(event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="not found")
    interest_counts = {}
    users_with_interests = 0
    total_votes = 0
    for u in _users_for(event_id):
        items = [s.strip() for s in (u.get("interest") or "").split(",") if s.strip()]
        if items:
            users_with_interests += 1
            for item in items:
                interest_counts[item] = interest_counts.get(item, 0) + 1
                total_votes += 1
    sorted_interests = sorted(interest_counts.items(), key=lambda x: x[1], reverse=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Interessefordeling"
    ws.append([f"Interessefordeling — {ev.get('name', '')}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"{users_with_interests} deltakere — {total_votes} stemmer totalt"])
    ws.append([])
    ws.append(["Interesseområde", "Antall stemmer", "Andel (%)", f"Av {total_votes}"])
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E0F0", end_color="E8E0F0", fill_type="solid")
    for interest, count in sorted_interests:
        pct = round(count / total_votes * 100, 1) if total_votes else 0
        ws.append([interest, count, pct, f"{count}/{total_votes}"])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=interessefordeling.xlsx"},
    )


# ── Admin: Booths (event-scoped) ──────────────────────────────────────────────

@app.get("/admin/api/events/{event_id}/booths")
def admin_list_event_booths(event_id: str, _=Depends(require_admin)):
    if not _event_by_id(event_id):
        raise HTTPException(status_code=404, detail="event not found")
    return _booths_for(event_id)


@app.post("/admin/api/events/{event_id}/booths")
async def admin_create_event_booth(event_id: str, _=Depends(require_admin)):
    if not _event_by_id(event_id):
        raise HTTPException(status_code=404, detail="event not found")
    pool = _booths_for(event_id)
    next_number = max((b.get("number", 0) for b in pool), default=0) + 1
    booth_id = str(uuid.uuid4())
    booth = {"id": booth_id, "name": f"Booth {next_number}", "number": next_number, "mode": "both", "event_id": event_id}
    booths.append(booth)
    try:
        await asyncio.get_event_loop().run_in_executor(
            None, instant_db.create_booth, booth_id, booth["name"], next_number, "both", event_id,
        )
    except Exception as e:
        print(f"[booth] create failed: {e}")
    return booth


@app.patch("/admin/api/events/{event_id}/booths/{booth_id}")
async def admin_update_event_booth(event_id: str, booth_id: str, body: dict, _=Depends(require_admin)):
    booth = next((b for b in _booths_for(event_id) if b["id"] == booth_id), None)
    if not booth:
        raise HTTPException(status_code=404, detail="not found")
    updates = {}
    mode = body.get("mode")
    if mode and mode in ("both", "register", "demo"):
        booth["mode"] = mode
        updates["mode"] = mode
    if "name" in body:
        booth["name"] = body["name"]
        updates["name"] = body["name"]
    if updates:
        try:
            await asyncio.get_event_loop().run_in_executor(
                None, lambda: instant_db.update_booth(booth_id, **updates),
            )
        except Exception as e:
            print(f"[booth] update failed: {e}")
    return {"ok": True, **booth}


@app.delete("/admin/api/events/{event_id}/booths/{booth_id}")
async def admin_delete_event_booth(event_id: str, booth_id: str, _=Depends(require_admin)):
    booth = next((b for b in _booths_for(event_id) if b["id"] == booth_id), None)
    if not booth:
        raise HTTPException(status_code=404, detail="not found")
    booths.remove(booth)
    try:
        await asyncio.get_event_loop().run_in_executor(None, instant_db.delete_booth, booth_id)
    except Exception as e:
        print(f"[booth] delete failed: {e}")
    return {"ok": True}


# Legacy admin user/booth routes → default event (for old admin.html if still hit)
@app.get("/admin/api/users")
def admin_list_users_legacy(_=Depends(require_admin)):
    d = _default_event()
    if not d:
        return []
    return admin_list_event_users(d["id"])


@app.get("/admin/api/booths")
def admin_list_booths_legacy(_=Depends(require_admin)):
    d = _default_event()
    if not d:
        return []
    return admin_list_event_booths(d["id"])


# ── Static: Admin React app + Booth SPA ───────────────────────────────────────

_STATIC_DIR = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=str(_STATIC_DIR)), name="static-files")

_WEB_DIR = Path(__file__).parent / "static" / "web"
_ADMIN_DIR = Path(__file__).parent / "static" / "admin-app"


@app.get("/admin", response_class=HTMLResponse)
def admin_page(_=Depends(require_admin)):
    index = _ADMIN_DIR / "index.html"
    if index.exists():
        return FileResponse(str(index))
    # Fallback to legacy admin during transition
    legacy = Path(__file__).parent / "static" / "admin.html"
    if legacy.exists():
        return HTMLResponse(legacy.read_text())
    return HTMLResponse("<h1>Admin app not built</h1><p>Run npm run build in web-admin/</p>", status_code=503)


if _ADMIN_DIR.exists():
    assets = _ADMIN_DIR / "assets"
    if assets.exists():
        app.mount("/admin/assets", StaticFiles(directory=str(assets)), name="admin-assets")

    @app.get("/admin/{path:path}")
    def admin_spa(path: str, _=Depends(require_admin)):
        if path.startswith("api/"):
            raise HTTPException(status_code=404, detail="not found")
        file_path = _ADMIN_DIR / path
        if file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(_ADMIN_DIR / "index.html"))


if _WEB_DIR.exists():
    app.mount("/booth/assets", StaticFiles(directory=str(_WEB_DIR / "assets")), name="web-assets")

    @app.get("/booth")
    @app.get("/booth/")
    def booth_index_redirect():
        d = _default_event()
        slug = d["slug"] if d else "lynskarp"
        return RedirectResponse(url=f"/e/{slug}/booth/1", status_code=302)

    @app.get("/booth/{path:path}")
    def booth_legacy_path(path: str):
        d = _default_event()
        slug = d["slug"] if d else "lynskarp"
        if path.isdigit():
            return RedirectResponse(url=f"/e/{slug}/booth/{path}", status_code=302)
        file_path = _WEB_DIR / path
        if file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(_WEB_DIR / "index.html"))


@app.get("/e/{slug}/booth")
@app.get("/e/{slug}/booth/")
def event_booth_index(slug: str):
    _require_event_slug(slug)
    if not _WEB_DIR.exists():
        raise HTTPException(status_code=503, detail="booth app not built")
    return FileResponse(str(_WEB_DIR / "index.html"))


@app.get("/e/{slug}/booth/assets/{asset_path:path}")
def event_booth_assets(slug: str, asset_path: str):
    _require_event_slug(slug)
    file_path = _WEB_DIR / "assets" / asset_path
    if file_path.exists() and file_path.is_file():
        return FileResponse(str(file_path))
    raise HTTPException(status_code=404, detail="not found")


@app.get("/e/{slug}/booth/{path:path}")
def event_booth_spa(slug: str, path: str):
    _require_event_slug(slug)
    if not _WEB_DIR.exists():
        raise HTTPException(status_code=503, detail="booth app not built")
    file_path = _WEB_DIR / path
    if file_path.exists() and file_path.is_file():
        return FileResponse(str(file_path))
    return FileResponse(str(_WEB_DIR / "index.html"))
